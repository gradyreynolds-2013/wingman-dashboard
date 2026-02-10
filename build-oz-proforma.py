#!/usr/bin/env python3
"""
OZ Fund Pro Forma Template Builder
Creates a professional Excel workbook for Opportunity Zone Fund real estate analysis
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import date, timedelta
import os

# Define styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow for inputs
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green for totals
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def format_header_row(ws, row, cols):
    """Apply header formatting to a row"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def format_section_row(ws, row, cols):
    """Apply section formatting to a row"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
        cell.border = THIN_BORDER

def format_input_cell(ws, row, col):
    """Apply input formatting to a cell"""
    cell = ws.cell(row=row, column=col)
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER

def format_total_row(ws, row, cols):
    """Apply total row formatting"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER

def set_column_widths(ws, widths):
    """Set column widths from a dictionary"""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def build_summary_dashboard(wb):
    """Sheet 1: Summary Dashboard"""
    ws = wb['1_Summary Dashboard']
    
    # Title
    ws['A1'] = "OPPORTUNITY ZONE FUND PRO FORMA"
    ws['A1'].font = Font(bold=True, size=18, color="4472C4")
    ws.merge_cells('A1:H1')
    
    ws['A2'] = "Multifamily Development Analysis"
    ws['A2'].font = Font(italic=True, size=12)
    ws.merge_cells('A2:H2')
    
    # PROJECT OVERVIEW SECTION
    ws['A4'] = "PROJECT OVERVIEW"
    format_section_row(ws, 4, 4)
    ws.merge_cells('A4:D4')
    
    labels = [
        ('A5', 'Project Name:', 'B5', 'OZ Multifamily Fund - Property 1'),
        ('A6', 'Location:', 'B6', 'Enter City, State'),
        ('A7', 'Property Type:', 'B7', 'Multifamily'),
        ('A8', 'Total Units:', 'B8', 50),
        ('A9', 'Year Built/Renovated:', 'B9', 2026),
        ('A10', 'OZ Designation:', 'B10', 'Qualified Opportunity Zone'),
    ]
    
    for label_cell, label, value_cell, value in labels:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = value
        if isinstance(value, str) and 'Enter' in value:
            format_input_cell(ws, int(value_cell[1:]), 2)
    
    # KEY METRICS BOX
    ws['F4'] = "KEY METRICS"
    format_section_row(ws, 4, 8)
    ws.merge_cells('F4:H4')
    
    metrics = [
        ('F5', 'Total Project Cost:', 'G5', "='2_Development Budget'!C32"),
        ('F6', 'Equity Required:', 'G6', "='5_Debt Schedule'!C7"),
        ('F7', 'Debt Amount:', 'G7', "='5_Debt Schedule'!C5"),
        ('F8', 'LTV %:', 'G8', "='5_Debt Schedule'!C4"),
        ('F9', 'Cap Rate (Going-In):', 'G9', "='4_Operating Expenses'!D18/'2_Development Budget'!C32"),
        ('F10', 'Price Per Unit:', 'G10', "='2_Development Budget'!C32/B8"),
    ]
    
    for label_cell, label, value_cell, formula in metrics:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = formula
        ws[value_cell].number_format = '$#,##0' if 'Cost' in label or 'Equity' in label or 'Debt' in label or 'Per Unit' in label else '0.00%'
    
    # RETURNS SNAPSHOT
    ws['A12'] = "RETURNS SNAPSHOT"
    format_section_row(ws, 12, 8)
    ws.merge_cells('A12:H12')
    
    returns = [
        ('A13', 'Year 1 Cash-on-Cash:', 'B13', "='6_Cash Flow Analysis'!C24"),
        ('A14', '10-Year IRR:', 'B14', "='7_Returns Exit'!C30"),
        ('A15', 'Equity Multiple:', 'B15', "='7_Returns Exit'!C29"),
        ('A16', 'Exit Value (Year 10):', 'B16', "='7_Returns Exit'!C9"),
        ('F13', 'Total Cash Returned:', 'G13', "='6_Cash Flow Analysis'!L25"),
        ('F14', 'Exit Equity Proceeds:', 'G14', "='7_Returns Exit'!C13"),
        ('F15', 'Total Return ($):', 'G15', "='7_Returns Exit'!C27"),
        ('F16', 'Total Return (%):', 'G16', "='7_Returns Exit'!C28"),
    ]
    
    for label_cell, label, value_cell, formula in returns:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = formula
        if '%' in label or 'IRR' in label or 'Multiple' in label:
            ws[value_cell].number_format = '0.00%'
        else:
            ws[value_cell].number_format = '$#,##0'
    
    # INVESTMENT TIMELINE
    ws['A18'] = "INVESTMENT TIMELINE"
    format_section_row(ws, 18, 8)
    ws.merge_cells('A18:H18')
    
    ws['A19'] = 'Year 0'
    ws['A20'] = 'Acquisition & Development'
    ws['B19'] = 'Years 1-2'
    ws['B20'] = 'Renovation & Lease-Up'
    ws['C19'] = 'Years 3-5'
    ws['C20'] = 'Stabilization & Refinance'
    ws['D19'] = 'Years 6-9'
    ws['D20'] = 'Cash Flow Operations'
    ws['E19'] = 'Year 10'
    ws['E20'] = 'Tax-Free Exit'
    
    for col in range(1, 6):
        ws.cell(row=19, column=col).font = Font(bold=True, color="4472C4")
        ws.cell(row=19, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=20, column=col).alignment = Alignment(horizontal='center')
    
    # OZ TAX BENEFITS
    ws['A22'] = "OZ TAX BENEFITS"
    format_section_row(ws, 22, 8)
    ws.merge_cells('A22:H22')
    
    tax_items = [
        ('A23', 'Original Capital Gain (Input):', 'B23', 1000000),
        ('A24', 'Federal Tax Rate:', 'B24', 0.238),
        ('A25', 'State Tax Rate:', 'B25', 0.05),
        ('A26', 'Tax Deferred (to Dec 2026):', 'B26', '=B23*(B24+B25)'),
        ('A27', 'Tax-Free Appreciation:', 'B27', "='7_Returns Exit'!C27"),
        ('F23', 'Traditional Investment Tax:', 'G23', "='7_Returns Exit'!C35"),
        ('F24', 'OZ Fund Tax:', 'G24', "='7_Returns Exit'!C39"),
        ('F25', 'Tax Savings from OZ:', 'G25', "='7_Returns Exit'!C42"),
    ]
    
    for label_cell, label, value_cell, value in tax_items:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = value
        if 'Rate' in label:
            ws[value_cell].number_format = '0.0%'
            format_input_cell(ws, int(value_cell[1:]), ord(value_cell[0])-64)
        elif 'Input' in label:
            ws[value_cell].number_format = '$#,##0'
            format_input_cell(ws, int(value_cell[1:]), ord(value_cell[0])-64)
        else:
            ws[value_cell].number_format = '$#,##0'
    
    # Column widths
    set_column_widths(ws, {'A': 25, 'B': 20, 'C': 20, 'D': 20, 'E': 20, 'F': 25, 'G': 18, 'H': 15})
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    return ws

def build_development_budget(wb):
    """Sheet 2: Development Budget"""
    ws = wb['2_Development Budget']
    
    # Header
    ws['A1'] = "DEVELOPMENT BUDGET"
    ws['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws.merge_cells('A1:D1')
    
    # Column headers
    headers = ['Category', 'Description', 'Amount', '% of Total']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    format_header_row(ws, 3, 4)
    
    row = 4
    
    # LAND ACQUISITION
    ws.cell(row=row, column=1, value='LAND ACQUISITION')
    format_section_row(ws, row, 4)
    
    land_items = [
        ('Land Purchase Price', 3000000),
        ('Closing Costs (2%)', '=C5*0.02'),
        ('Due Diligence', 25000),
    ]
    
    land_start = row + 1
    for item, value in land_items:
        row += 1
        ws.cell(row=row, column=2, value=item)
        ws.cell(row=row, column=3, value=value)
        if isinstance(value, (int, float)):
            format_input_cell(ws, row, 3)
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=f'=C{row}/$C$32')
        ws.cell(row=row, column=4).number_format = '0.0%'
    
    row += 1
    land_end = row - 1
    ws.cell(row=row, column=2, value='Subtotal Land')
    ws.cell(row=row, column=3, value=f'=SUM(C{land_start}:C{land_end})')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    ws.cell(row=row, column=4, value=f'=C{row}/$C$32')
    ws.cell(row=row, column=4).number_format = '0.0%'
    format_total_row(ws, row, 4)
    land_subtotal_row = row  # Row 8
    
    # HARD CONSTRUCTION COSTS
    row += 1
    ws.cell(row=row, column=1, value='HARD CONSTRUCTION COSTS')
    format_section_row(ws, row, 4)
    
    hard_items = [
        ('Site Work & Utilities', 500000),
        ('Building Construction', 8500000),
        ('Parking & Landscaping', 400000),
        ('FF&E (Common Areas)', 300000),
        ('Unit Finishes/Upgrades', 300000),
    ]
    
    hard_start = row + 1
    for item, value in hard_items:
        row += 1
        ws.cell(row=row, column=2, value=item)
        ws.cell(row=row, column=3, value=value)
        format_input_cell(ws, row, 3)
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=f'=C{row}/$C$32')
        ws.cell(row=row, column=4).number_format = '0.0%'
    
    row += 1
    hard_end = row - 1
    ws.cell(row=row, column=2, value='Subtotal Hard Costs')
    ws.cell(row=row, column=3, value=f'=SUM(C{hard_start}:C{hard_end})')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    ws.cell(row=row, column=4, value=f'=C{row}/$C$32')
    ws.cell(row=row, column=4).number_format = '0.0%'
    format_total_row(ws, row, 4)
    hard_subtotal_row = row  # Row 15
    
    # SOFT COSTS
    row += 1
    ws.cell(row=row, column=1, value='SOFT COSTS')
    format_section_row(ws, row, 4)
    
    soft_items = [
        ('Architecture & Design', 350000),
        ('Engineering', 150000),
        ('Permits & Impact Fees', 200000),
        ('Legal & Accounting', 100000),
        ('Financing Fees (2% of loan)', f'=$C$32*0.75*0.02'),
        ('Construction Interest', 250000),
        ('Insurance (Builder\'s Risk)', 75000),
        ('Marketing & Lease-Up', 100000),
    ]
    
    soft_start = row + 1
    for item, value in soft_items:
        row += 1
        ws.cell(row=row, column=2, value=item)
        ws.cell(row=row, column=3, value=value)
        if isinstance(value, (int, float)):
            format_input_cell(ws, row, 3)
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=f'=C{row}/$C$32')
        ws.cell(row=row, column=4).number_format = '0.0%'
    
    row += 1
    soft_end = row - 1
    ws.cell(row=row, column=2, value='Subtotal Soft Costs')
    ws.cell(row=row, column=3, value=f'=SUM(C{soft_start}:C{soft_end})')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    ws.cell(row=row, column=4, value=f'=C{row}/$C$32')
    ws.cell(row=row, column=4).number_format = '0.0%'
    format_total_row(ws, row, 4)
    soft_subtotal_row = row  # Row 25
    
    # DEVELOPMENT FEE
    row += 1
    ws.cell(row=row, column=1, value='DEVELOPMENT FEE')
    format_section_row(ws, row, 4)
    
    row += 1
    ws.cell(row=row, column=2, value='Development Fee (5% of hard costs)')
    ws.cell(row=row, column=3, value=f'=C{hard_subtotal_row}*0.05')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    ws.cell(row=row, column=4, value=f'=C{row}/$C$32')
    ws.cell(row=row, column=4).number_format = '0.0%'
    dev_fee_row = row  # Row 27
    
    # CONTINGENCY
    row += 1
    ws.cell(row=row, column=1, value='CONTINGENCY')
    format_section_row(ws, row, 4)
    
    row += 1
    ws.cell(row=row, column=2, value='Contingency (7% of hard + soft)')
    ws.cell(row=row, column=3, value=f'=(C{hard_subtotal_row}+C{soft_subtotal_row})*0.07')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    ws.cell(row=row, column=4, value=f'=C{row}/$C$32')
    ws.cell(row=row, column=4).number_format = '0.0%'
    contingency_row = row  # Row 29
    
    # TOTAL PROJECT COST
    row += 2
    ws.cell(row=row, column=1, value='TOTAL PROJECT COST')
    ws.cell(row=row, column=3, value=f'=C{land_subtotal_row}+C{hard_subtotal_row}+C{soft_subtotal_row}+C{dev_fee_row}+C{contingency_row}')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    ws.cell(row=row, column=4, value='100.0%')
    format_total_row(ws, row, 4)
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=3).font = Font(bold=True, size=12)
    total_cost_row = row  # Row 32
    
    # Summary metrics
    row += 2
    ws.cell(row=row, column=1, value='Cost Per Unit:')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f"=C{total_cost_row}/'1_Summary Dashboard'!B8")
    ws.cell(row=row, column=3).number_format = '$#,##0'
    
    row += 1
    ws.cell(row=row, column=1, value='Hard Cost Per Unit:')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f"=C{hard_subtotal_row}/'1_Summary Dashboard'!B8")
    ws.cell(row=row, column=3).number_format = '$#,##0'
    
    # Column widths
    set_column_widths(ws, {'A': 30, 'B': 35, 'C': 18, 'D': 12})
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    return ws

def build_unit_mix_revenue(wb):
    """Sheet 3: Unit Mix & Revenue"""
    ws = wb['3_Unit Mix Revenue']
    
    # Header
    ws['A1'] = "UNIT MIX & REVENUE PROJECTIONS"
    ws['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws.merge_cells('A1:N1')
    
    # Column headers - Years 1-10
    headers = ['Unit Type', 'Units', 'Mo. Rent', 'Growth']
    years = ['Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', 'Year 6', 'Year 7', 'Year 8', 'Year 9', 'Year 10']
    all_headers = headers + years
    
    for col, header in enumerate(all_headers, 1):
        ws.cell(row=3, column=col, value=header)
    format_header_row(ws, 3, len(all_headers))
    
    # UNIT MIX
    ws['A4'] = 'RENTAL INCOME'
    format_section_row(ws, 4, len(all_headers))
    
    # Unit types with data
    unit_data = [
        ('Studio', 10, 1200, 0.03),
        ('1 Bedroom', 25, 1500, 0.03),
        ('2 Bedroom', 12, 2000, 0.03),
        ('3 Bedroom', 3, 2400, 0.03),
    ]
    
    row = 4
    unit_start = row + 1
    for unit_type, units, rent, growth in unit_data:
        row += 1
        ws.cell(row=row, column=1, value=unit_type)
        ws.cell(row=row, column=2, value=units)
        format_input_cell(ws, row, 2)
        ws.cell(row=row, column=3, value=rent)
        format_input_cell(ws, row, 3)
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=growth)
        format_input_cell(ws, row, 4)
        ws.cell(row=row, column=4).number_format = '0.0%'
        
        # Year calculations (annual rent = units * monthly rent * 12)
        # Year 1
        ws.cell(row=row, column=5, value=f'=B{row}*C{row}*12')
        ws.cell(row=row, column=5).number_format = '$#,##0'
        
        # Years 2-10 (with growth)
        for year in range(2, 11):
            col = year + 4
            prev_col = get_column_letter(col - 1)
            ws.cell(row=row, column=col, value=f'={prev_col}{row}*(1+$D{row})')
            ws.cell(row=row, column=col).number_format = '$#,##0'
    
    unit_end = row  # Row 8
    
    # Total Units row
    row += 1
    ws.cell(row=row, column=1, value='Total Units')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=f'=SUM(B{unit_start}:B{unit_end})')
    total_units_row = row  # Row 9
    
    # GROSS POTENTIAL RENT
    row += 1
    ws.cell(row=row, column=1, value='GROSS POTENTIAL RENT')
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col in range(5, 15):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{unit_start}:{col_letter}{unit_end})')
        ws.cell(row=row, column=col).number_format = '$#,##0'
    format_total_row(ws, row, len(all_headers))
    gpr_row = row  # Row 10
    
    # VACANCY & ADJUSTMENTS
    row += 2
    ws.cell(row=row, column=1, value='VACANCY & ADJUSTMENTS')
    format_section_row(ws, row, len(all_headers))
    
    vacancy_items = [
        ('Physical Vacancy', 0.05),
        ('Collection Loss/Bad Debt', 0.005),
        ('Concessions', 0.005),
    ]
    
    vac_start = row + 1
    for item, rate in vacancy_items:
        row += 1
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=rate)
        format_input_cell(ws, row, 2)
        ws.cell(row=row, column=2).number_format = '0.0%'
        
        for col in range(5, 15):
            col_letter = get_column_letter(col)
            ws.cell(row=row, column=col, value=f'=-{col_letter}${gpr_row}*$B{row}')
            ws.cell(row=row, column=col).number_format = '($#,##0)'
    
    vac_end = row  # Row 15
    
    # Total Vacancy Loss
    row += 1
    ws.cell(row=row, column=1, value='Total Vacancy/Loss')
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col in range(5, 15):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{vac_start}:{col_letter}{vac_end})')
        ws.cell(row=row, column=col).number_format = '($#,##0)'
    vacancy_total_row = row  # Row 16
    
    # GROSS OPERATING INCOME
    row += 1
    ws.cell(row=row, column=1, value='GROSS OPERATING INCOME')
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col in range(5, 15):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={col_letter}{gpr_row}+{col_letter}{vacancy_total_row}')
        ws.cell(row=row, column=col).number_format = '$#,##0'
    format_total_row(ws, row, len(all_headers))
    goi_row = row  # Row 17
    
    # OTHER INCOME
    row += 2
    ws.cell(row=row, column=1, value='OTHER INCOME')
    format_section_row(ws, row, len(all_headers))
    
    other_income = [
        ('Parking', 50, 0.03),
        ('Storage', 25, 0.03),
        ('Laundry', 15, 0.03),
        ('Pet Fees', 30, 0.03),
        ('Application/Admin Fees', 10, 0.03),
    ]
    
    other_start = row + 1
    for item, per_unit_monthly, growth in other_income:
        row += 1
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=per_unit_monthly)
        format_input_cell(ws, row, 2)
        ws.cell(row=row, column=2).number_format = '$#,##0'
        ws.cell(row=row, column=3, value='per unit/mo')
        ws.cell(row=row, column=4, value=growth)
        format_input_cell(ws, row, 4)
        ws.cell(row=row, column=4).number_format = '0.0%'
        
        # Year 1
        ws.cell(row=row, column=5, value=f'=$B${total_units_row}*$B{row}*12')
        ws.cell(row=row, column=5).number_format = '$#,##0'
        
        # Years 2-10
        for year in range(2, 11):
            col = year + 4
            prev_col = get_column_letter(col - 1)
            ws.cell(row=row, column=col, value=f'={prev_col}{row}*(1+$D{row})')
            ws.cell(row=row, column=col).number_format = '$#,##0'
    
    other_end = row  # Row 24
    
    # Total Other Income
    row += 1
    ws.cell(row=row, column=1, value='Total Other Income')
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col in range(5, 15):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{other_start}:{col_letter}{other_end})')
        ws.cell(row=row, column=col).number_format = '$#,##0'
    format_total_row(ws, row, len(all_headers))
    other_total_row = row  # Row 25
    
    # EFFECTIVE GROSS INCOME
    row += 2
    ws.cell(row=row, column=1, value='EFFECTIVE GROSS INCOME (EGI)')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    for col in range(5, 15):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={col_letter}{goi_row}+{col_letter}{other_total_row}')
        ws.cell(row=row, column=col).number_format = '$#,##0'
        ws.cell(row=row, column=col).font = Font(bold=True)
    format_total_row(ws, row, len(all_headers))
    # EGI row = 27
    
    # Column widths
    set_column_widths(ws, {'A': 25, 'B': 10, 'C': 12, 'D': 10})
    for i in range(5, 15):
        ws.column_dimensions[get_column_letter(i)].width = 12
    
    # Freeze panes
    ws.freeze_panes = 'E4'
    
    return ws

def build_operating_expenses(wb):
    """Sheet 4: Operating Expenses"""
    ws = wb['4_Operating Expenses']
    
    # Header
    ws['A1'] = "OPERATING EXPENSES"
    ws['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws.merge_cells('A1:M1')
    
    # Column headers
    headers = ['Expense Category', '% of EGI', 'Growth']
    years = ['Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', 'Year 6', 'Year 7', 'Year 8', 'Year 9', 'Year 10']
    all_headers = headers + years
    
    for col, header in enumerate(all_headers, 1):
        ws.cell(row=3, column=col, value=header)
    format_header_row(ws, 3, len(all_headers))
    
    # Reference EGI from Unit Mix sheet (row 27)
    row = 4
    ws.cell(row=row, column=1, value='Effective Gross Income (Ref)')
    ws.cell(row=row, column=1).font = Font(italic=True)
    for year in range(1, 11):
        col = year + 3  # D=4, E=5, etc.
        rev_col_letter = get_column_letter(year + 4)  # Revenue sheet year cols start at E
        ws.cell(row=row, column=col, value=f"='3_Unit Mix Revenue'!{rev_col_letter}27")
        ws.cell(row=row, column=col).number_format = '$#,##0'
        ws.cell(row=row, column=col).font = Font(italic=True)
    egi_ref_row = row  # Row 4
    
    # OPERATING EXPENSES
    row += 1
    ws.cell(row=row, column=1, value='OPERATING EXPENSES')
    format_section_row(ws, row, len(all_headers))
    
    # Expense items with % of EGI and growth rate
    expenses = [
        ('Payroll', 0.05, 0.03),
        ('Utilities', 0.03, 0.025),
        ('Repairs & Maintenance', 0.04, 0.03),
        ('Property Taxes', 0.06, 0.02),
        ('Insurance', 0.02, 0.04),
        ('Marketing & Leasing', 0.015, 0.03),
        ('General & Administrative', 0.02, 0.03),
        ('Management Fee', 0.03, 0.00),  # % of EGI, no growth (tracks EGI)
        ('Replacement Reserves', 0.02, 0.03),
    ]
    
    exp_start = row + 1
    for expense, pct, growth in expenses:
        row += 1
        ws.cell(row=row, column=1, value=expense)
        ws.cell(row=row, column=2, value=pct)
        format_input_cell(ws, row, 2)
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=3, value=growth)
        format_input_cell(ws, row, 3)
        ws.cell(row=row, column=3).number_format = '0.0%'
        
        if expense == 'Management Fee':
            # Management fee tracks EGI directly
            for col in range(4, 14):
                col_letter = get_column_letter(col)
                ws.cell(row=row, column=col, value=f'=${col_letter}${egi_ref_row}*$B{row}')
                ws.cell(row=row, column=col).number_format = '$#,##0'
        else:
            # Year 1 based on % of EGI
            ws.cell(row=row, column=4, value=f'=$D${egi_ref_row}*$B{row}')
            ws.cell(row=row, column=4).number_format = '$#,##0'
            
            # Years 2-10 with growth
            for col in range(5, 14):
                prev_col = get_column_letter(col - 1)
                ws.cell(row=row, column=col, value=f'={prev_col}{row}*(1+$C{row})')
                ws.cell(row=row, column=col).number_format = '$#,##0'
    
    exp_end = row  # Row 14
    
    # TOTAL OPERATING EXPENSES
    row += 1
    ws.cell(row=row, column=1, value='TOTAL OPERATING EXPENSES')
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col in range(4, 14):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=SUM({col_letter}{exp_start}:{col_letter}{exp_end})')
        ws.cell(row=row, column=col).number_format = '$#,##0'
    format_total_row(ws, row, len(all_headers))
    total_exp_row = row  # Row 15
    
    # Operating Expense Ratio
    row += 1
    ws.cell(row=row, column=1, value='Operating Expense Ratio')
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col in range(4, 14):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={col_letter}{total_exp_row}/{col_letter}{egi_ref_row}')
        ws.cell(row=row, column=col).number_format = '0.0%'
    exp_ratio_row = row  # Row 16
    
    # NET OPERATING INCOME
    row += 2
    ws.cell(row=row, column=1, value='NET OPERATING INCOME (NOI)')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    for col in range(4, 14):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={col_letter}{egi_ref_row}-{col_letter}{total_exp_row}')
        ws.cell(row=row, column=col).number_format = '$#,##0'
        ws.cell(row=row, column=col).font = Font(bold=True)
    format_total_row(ws, row, len(all_headers))
    # NOI row = 18
    
    # Column widths
    set_column_widths(ws, {'A': 28, 'B': 12, 'C': 10})
    for i in range(4, 14):
        ws.column_dimensions[get_column_letter(i)].width = 12
    
    # Freeze panes
    ws.freeze_panes = 'D4'
    
    return ws

def build_debt_schedule(wb):
    """Sheet 5: Debt Schedule"""
    ws = wb['5_Debt Schedule']
    
    # Header
    ws['A1'] = "DEBT SCHEDULE"
    ws['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws.merge_cells('A1:G1')
    
    # LOAN INPUTS
    ws['A3'] = 'LOAN INPUTS'
    format_section_row(ws, 3, 4)
    
    row = 3
    
    row += 1  # Row 4
    ws.cell(row=row, column=1, value='Loan-to-Value (LTV)')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=0.75)
    ws.cell(row=row, column=3).number_format = '0.0%'
    format_input_cell(ws, row, 3)
    ltv_row = row
    
    row += 1  # Row 5
    ws.cell(row=row, column=1, value='Loan Amount')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value="='2_Development Budget'!C32*C4")
    ws.cell(row=row, column=3).number_format = '$#,##0'
    loan_amt_row = row
    
    row += 1  # Row 6
    ws.cell(row=row, column=1, value='Interest Rate')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=0.065)
    ws.cell(row=row, column=3).number_format = '0.00%'
    format_input_cell(ws, row, 3)
    int_rate_row = row
    
    row += 1  # Row 7
    ws.cell(row=row, column=1, value='Equity Required')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value="='2_Development Budget'!C32-C5")
    ws.cell(row=row, column=3).number_format = '$#,##0'
    format_total_row(ws, row, 4)
    equity_row = row
    
    row += 1  # Row 8
    ws.cell(row=row, column=1, value='Amortization (Years)')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=30)
    format_input_cell(ws, row, 3)
    amort_row = row
    
    row += 1  # Row 9
    ws.cell(row=row, column=1, value='Loan Term (Years)')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=10)
    format_input_cell(ws, row, 3)
    term_row = row
    
    # Annual Debt Service
    row += 1  # Row 10
    ws.cell(row=row, column=1, value='Annual Debt Service')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f'=PMT(C{int_rate_row},C{amort_row},-C{loan_amt_row})*12')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    format_total_row(ws, row, 4)
    debt_svc_row = row
    
    # AMORTIZATION SCHEDULE
    row += 2  # Row 12
    ws.cell(row=row, column=1, value='AMORTIZATION SCHEDULE')
    format_section_row(ws, row, 7)
    
    row += 1  # Row 13
    amort_headers = ['Year', 'Beginning Balance', 'Annual Payment', 'Interest', 'Principal', 'Ending Balance', 'DCR']
    for col, header in enumerate(amort_headers, 1):
        ws.cell(row=row, column=col, value=header)
    format_header_row(ws, row, len(amort_headers))
    header_row = row
    
    # Years 1-10
    for year in range(1, 11):
        row += 1
        ws.cell(row=row, column=1, value=year)
        
        if year == 1:
            ws.cell(row=row, column=2, value=f'=$C${loan_amt_row}')  # Beginning balance = loan amount
        else:
            ws.cell(row=row, column=2, value=f'=F{row-1}')  # Beginning = previous ending
        ws.cell(row=row, column=2).number_format = '$#,##0'
        
        # Annual Payment
        ws.cell(row=row, column=3, value=f'=$C${debt_svc_row}')
        ws.cell(row=row, column=3).number_format = '$#,##0'
        
        # Interest
        ws.cell(row=row, column=4, value=f'=B{row}*$C${int_rate_row}')
        ws.cell(row=row, column=4).number_format = '$#,##0'
        
        # Principal
        ws.cell(row=row, column=5, value=f'=C{row}-D{row}')
        ws.cell(row=row, column=5).number_format = '$#,##0'
        
        # Ending Balance
        ws.cell(row=row, column=6, value=f'=B{row}-E{row}')
        ws.cell(row=row, column=6).number_format = '$#,##0'
        
        # DCR (NOI / Debt Service) - NOI is in row 18 of Operating Expenses, cols D-M
        noi_col = get_column_letter(year + 3)  # D=Year1, E=Year2, etc.
        ws.cell(row=row, column=7, value=f"='4_Operating Expenses'!{noi_col}18/$C${debt_svc_row}")
        ws.cell(row=row, column=7).number_format = '0.00x'
    
    # Column widths
    set_column_widths(ws, {'A': 20, 'B': 18, 'C': 16, 'D': 14, 'E': 14, 'F': 16, 'G': 10})
    
    # Freeze panes
    ws.freeze_panes = 'A14'
    
    return ws

def build_cash_flow_analysis(wb):
    """Sheet 6: Cash Flow Analysis"""
    ws = wb['6_Cash Flow Analysis']
    
    # Header
    ws['A1'] = "CASH FLOW ANALYSIS"
    ws['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws.merge_cells('A1:L1')
    
    # Column headers - No Year 0, just Years 1-10 + Total
    headers = ['Cash Flow Item', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', 'Year 6', 'Year 7', 'Year 8', 'Year 9', 'Year 10', 'Total']
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    format_header_row(ws, 3, len(headers))
    
    # CASH FLOWS
    ws['A4'] = 'OPERATING CASH FLOWS'
    format_section_row(ws, 4, len(headers))
    
    # Net Operating Income - links to Operating Expenses sheet row 18, cols D-M
    row = 5
    ws.cell(row=row, column=1, value='Net Operating Income')
    for year in range(1, 11):
        col = year + 1  # B=2, C=3, etc.
        noi_col = get_column_letter(year + 3)  # D=Year1, E=Year2, etc.
        ws.cell(row=row, column=col, value=f"='4_Operating Expenses'!{noi_col}18")
        ws.cell(row=row, column=col).number_format = '$#,##0'
    ws.cell(row=row, column=12, value='=SUM(B5:K5)')
    ws.cell(row=row, column=12).number_format = '$#,##0'
    noi_row = row
    
    # Debt Service - links to Debt Schedule row 10
    row = 6
    ws.cell(row=row, column=1, value='Less: Debt Service')
    for col in range(2, 12):
        ws.cell(row=row, column=col, value="=-'5_Debt Schedule'!$C$10")
        ws.cell(row=row, column=col).number_format = '($#,##0)'
    ws.cell(row=row, column=12, value='=SUM(B6:K6)')
    ws.cell(row=row, column=12).number_format = '($#,##0)'
    debt_svc_row = row
    
    # Free Cash Flow (before refinancing)
    row = 7
    ws.cell(row=row, column=1, value='Free Cash Flow (Pre-Refi)')
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col in range(2, 13):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={col_letter}{noi_row}+{col_letter}{debt_svc_row}')
        ws.cell(row=row, column=col).number_format = '$#,##0'
    format_total_row(ws, row, len(headers))
    fcf_row = row
    
    # REFINANCING
    row = 9
    ws['A9'] = 'REFINANCING (Year 4)'
    format_section_row(ws, 9, len(headers))
    
    # Refinancing inputs
    row = 10
    ws.cell(row=row, column=1, value='Refi LTV %')
    ws.cell(row=row, column=2, value=0.75)
    format_input_cell(ws, row, 2)
    ws.cell(row=row, column=2).number_format = '0.0%'
    refi_ltv_row = row
    
    row = 11
    ws.cell(row=row, column=1, value='Refi Cap Rate')
    ws.cell(row=row, column=2, value=0.055)
    format_input_cell(ws, row, 2)
    ws.cell(row=row, column=2).number_format = '0.00%'
    refi_cap_row = row
    
    # Stabilized Value (NOI / Cap Rate) - Use Year 4 NOI (col E in cash flow = col G in expenses)
    row = 12
    ws.cell(row=row, column=1, value='Stabilized Value at Refi')
    ws.cell(row=row, column=5, value=f'=E{noi_row}/$B${refi_cap_row}')  # Year 4 = col E
    ws.cell(row=row, column=5).number_format = '$#,##0'
    stab_val_row = row
    
    # New Loan Amount
    row = 13
    ws.cell(row=row, column=1, value='New Loan Amount (75% LTV)')
    ws.cell(row=row, column=5, value=f'=E{stab_val_row}*$B${refi_ltv_row}')
    ws.cell(row=row, column=5).number_format = '$#,##0'
    new_loan_row = row
    
    # Existing Loan Balance (from debt schedule) - Year 4 ending balance is row 17, col F
    row = 14
    ws.cell(row=row, column=1, value='Less: Existing Loan Balance')
    ws.cell(row=row, column=5, value="=-'5_Debt Schedule'!F17")
    ws.cell(row=row, column=5).number_format = '($#,##0)'
    exist_loan_row = row
    
    # Refinancing Proceeds
    row = 15
    ws.cell(row=row, column=1, value='Refinancing Proceeds')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=5, value=f'=E{new_loan_row}+E{exist_loan_row}')  # New loan - existing balance
    ws.cell(row=row, column=5).number_format = '$#,##0'
    format_total_row(ws, 15, len(headers))
    ws.cell(row=row, column=12, value='=E15')
    ws.cell(row=row, column=12).number_format = '$#,##0'
    refi_proceeds_row = row
    
    # INVESTOR CASH FLOWS
    row = 17
    ws['A17'] = 'INVESTOR CASH FLOWS'
    format_section_row(ws, 17, len(headers))
    
    # Initial Investment info
    row = 18
    ws.cell(row=row, column=1, value='Initial Equity (for reference)')
    ws.cell(row=row, column=2, value="='5_Debt Schedule'!C7")
    ws.cell(row=row, column=2).number_format = '$#,##0'
    init_equity_row = row
    
    # Asset Management Fee (3% of free cash flow)
    row = 19
    ws.cell(row=row, column=1, value='Less: Asset Mgmt Fee (3%)')
    for col in range(2, 12):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'=-{col_letter}{fcf_row}*0.03')
        ws.cell(row=row, column=col).number_format = '($#,##0)'
    ws.cell(row=row, column=12, value='=SUM(B19:K19)')
    ws.cell(row=row, column=12).number_format = '($#,##0)'
    asset_mgmt_row = row
    
    # Net Cash Flow to Investors
    row = 20
    ws.cell(row=row, column=1, value='Net Cash Flow to Investors')
    ws.cell(row=row, column=1).font = Font(bold=True)
    for col in range(2, 12):
        col_letter = get_column_letter(col)
        # Free cash flow + refi proceeds (if Year 4, col E) + asset mgmt fee
        if col == 5:  # Year 4 with refi (column E)
            ws.cell(row=row, column=col, value=f'={col_letter}{fcf_row}+{col_letter}{refi_proceeds_row}+{col_letter}{asset_mgmt_row}')
        else:
            ws.cell(row=row, column=col, value=f'={col_letter}{fcf_row}+{col_letter}{asset_mgmt_row}')
        ws.cell(row=row, column=col).number_format = '$#,##0'
    ws.cell(row=row, column=12, value='=SUM(B20:K20)')
    ws.cell(row=row, column=12).number_format = '$#,##0'
    format_total_row(ws, row, len(headers))
    net_cf_row = row
    
    # RETURNS METRICS
    row = 22
    ws['A22'] = 'RETURNS METRICS'
    format_section_row(ws, 22, len(headers))
    
    # Cash-on-Cash Return
    row = 24
    ws.cell(row=row, column=1, value='Cash-on-Cash Return')
    for col in range(2, 12):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={col_letter}{net_cf_row}/$B${init_equity_row}')
        ws.cell(row=row, column=col).number_format = '0.00%'
    coc_row = row
    
    # Cumulative Cash Returned
    row = 25
    ws.cell(row=row, column=1, value='Cumulative Cash Returned')
    ws.cell(row=row, column=2, value=f'=B{net_cf_row}')
    ws.cell(row=row, column=2).number_format = '$#,##0'
    for col in range(3, 12):
        col_letter = get_column_letter(col)
        prev_col = get_column_letter(col - 1)
        ws.cell(row=row, column=col, value=f'={prev_col}{row}+{col_letter}{net_cf_row}')
        ws.cell(row=row, column=col).number_format = '$#,##0'
    cumulative_row = row
    
    # % of Original Equity Returned
    row = 26
    ws.cell(row=row, column=1, value='% of Equity Returned')
    for col in range(2, 12):
        col_letter = get_column_letter(col)
        ws.cell(row=row, column=col, value=f'={col_letter}{cumulative_row}/$B${init_equity_row}')
        ws.cell(row=row, column=col).number_format = '0.0%'
    
    # Column widths
    set_column_widths(ws, {'A': 28})
    for i in range(2, 13):
        ws.column_dimensions[get_column_letter(i)].width = 12
    
    # Freeze panes
    ws.freeze_panes = 'B4'
    
    return ws

def build_returns_exit(wb):
    """Sheet 7: Returns & Exit Analysis"""
    ws = wb['7_Returns Exit']
    
    # Header
    ws['A1'] = "RETURNS & EXIT ANALYSIS"
    ws['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws.merge_cells('A1:D1')
    
    # EXIT ASSUMPTIONS
    ws['A3'] = 'EXIT ASSUMPTIONS (YEAR 10)'
    format_section_row(ws, 3, 4)
    
    row = 3
    
    row += 1  # Row 4
    ws.cell(row=row, column=1, value='Exit Cap Rate')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=0.055)
    ws.cell(row=row, column=3).number_format = '0.00%'
    format_input_cell(ws, row, 3)
    exit_cap_row = row
    
    row += 1  # Row 5
    ws.cell(row=row, column=1, value='Selling Costs')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=0.02)
    ws.cell(row=row, column=3).number_format = '0.0%'
    format_input_cell(ws, row, 3)
    sell_cost_row = row
    
    # EXIT CALCULATIONS
    row += 2  # Row 7
    ws.cell(row=row, column=1, value='EXIT CALCULATIONS')
    format_section_row(ws, row, 4)
    
    row += 1  # Row 8
    ws.cell(row=row, column=1, value='Projected NOI (Year 10)')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value="='4_Operating Expenses'!M18")  # Year 10 NOI
    ws.cell(row=row, column=3).number_format = '$#,##0'
    noi_y10_row = row
    
    row += 1  # Row 9
    ws.cell(row=row, column=1, value='Gross Sale Price (NOI / Cap)')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f'=C{noi_y10_row}/C{exit_cap_row}')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    gross_sale_row = row
    
    row += 1  # Row 10
    ws.cell(row=row, column=1, value='Less: Selling Costs')
    ws.cell(row=row, column=3, value=f'=-C{gross_sale_row}*C{sell_cost_row}')
    ws.cell(row=row, column=3).number_format = '($#,##0)'
    
    row += 1  # Row 11
    ws.cell(row=row, column=1, value='Net Sale Proceeds')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f'=C{gross_sale_row}+C{row-1}')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    net_sale_row = row
    
    row += 1  # Row 12
    ws.cell(row=row, column=1, value='Less: Remaining Loan Balance')
    ws.cell(row=row, column=3, value="=-'5_Debt Schedule'!F23")  # Year 10 ending balance
    ws.cell(row=row, column=3).number_format = '($#,##0)'
    
    row += 1  # Row 13
    ws.cell(row=row, column=1, value='Equity Proceeds at Sale')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f'=C{net_sale_row}+C{row-1}')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    format_total_row(ws, row, 4)
    equity_proceeds_row = row
    
    # TOTAL RETURN CALCULATION
    row += 2  # Row 15
    ws.cell(row=row, column=1, value='TOTAL RETURN CALCULATION')
    format_section_row(ws, row, 4)
    
    row += 1  # Row 16
    ws.cell(row=row, column=1, value='Initial Equity Investment')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value="='5_Debt Schedule'!C7")
    ws.cell(row=row, column=3).number_format = '$#,##0'
    init_equity_row = row
    
    row += 1  # Row 17
    ws.cell(row=row, column=1, value='Total Operating Cash Flow')
    ws.cell(row=row, column=3, value="='6_Cash Flow Analysis'!L7")
    ws.cell(row=row, column=3).number_format = '$#,##0'
    total_op_cf_row = row
    
    row += 1  # Row 18
    ws.cell(row=row, column=1, value='Less: Asset Management Fees')
    ws.cell(row=row, column=3, value="='6_Cash Flow Analysis'!L19")
    ws.cell(row=row, column=3).number_format = '($#,##0)'
    asset_fee_row = row
    
    row += 1  # Row 19
    ws.cell(row=row, column=1, value='Refinancing Proceeds')
    ws.cell(row=row, column=3, value="='6_Cash Flow Analysis'!L15")
    ws.cell(row=row, column=3).number_format = '$#,##0'
    refi_row = row
    
    row += 1  # Row 20
    ws.cell(row=row, column=1, value='Net Operating Distributions')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f'=C{total_op_cf_row}+C{asset_fee_row}+C{refi_row}')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    net_dist_row = row
    
    row += 1  # Row 21
    ws.cell(row=row, column=1, value='Plus: Equity Proceeds at Exit')
    ws.cell(row=row, column=3, value=f'=C{equity_proceeds_row}')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    exit_equity_row = row
    
    row += 2  # Row 23
    ws.cell(row=row, column=1, value='TOTAL RETURNS')
    format_section_row(ws, row, 4)
    
    row += 1  # Row 24
    ws.cell(row=row, column=1, value='Gross Proceeds')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f'=C{net_dist_row}+C{exit_equity_row}')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    gross_proceeds_row = row
    
    row += 1  # Row 25
    ws.cell(row=row, column=1, value='Less: Initial Investment')
    ws.cell(row=row, column=3, value=f'=-C{init_equity_row}')
    ws.cell(row=row, column=3).number_format = '($#,##0)'
    
    row += 2  # Row 27
    ws.cell(row=row, column=1, value='Total Return ($)')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value=f'=C{gross_proceeds_row}-C{init_equity_row}')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    format_total_row(ws, row, 4)
    total_return_row = row
    
    row += 1  # Row 28
    ws.cell(row=row, column=1, value='Total Return (%)')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value=f'=C{total_return_row}/C{init_equity_row}')
    ws.cell(row=row, column=3).number_format = '0.0%'
    format_total_row(ws, row, 4)
    
    row += 1  # Row 29
    ws.cell(row=row, column=1, value='Equity Multiple')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row, column=3, value=f'=C{gross_proceeds_row}/C{init_equity_row}')
    ws.cell(row=row, column=3).number_format = '0.00x'
    format_total_row(ws, row, 4)
    
    row += 1  # Row 30
    ws.cell(row=row, column=1, value='Approx. Annualized IRR')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    # Simplified IRR approximation: (Multiple^(1/years))-1
    ws.cell(row=row, column=3, value=f'=(C{row-1}^(1/10))-1')
    ws.cell(row=row, column=3).number_format = '0.00%'
    format_total_row(ws, row, 4)
    
    # OZ TAX BENEFIT ANALYSIS
    row += 2  # Row 32
    ws.cell(row=row, column=1, value='OZ TAX BENEFIT ANALYSIS')
    format_section_row(ws, row, 4)
    
    row += 1  # Row 33
    ws.cell(row=row, column=1, value='Capital Gain Invested')
    ws.cell(row=row, column=3, value="='1_Summary Dashboard'!B23")
    ws.cell(row=row, column=3).number_format = '$#,##0'
    cap_gain_row = row
    
    row += 1  # Row 34
    ws.cell(row=row, column=1, value='Combined Tax Rate')
    ws.cell(row=row, column=3, value="='1_Summary Dashboard'!B24+'1_Summary Dashboard'!B25")
    ws.cell(row=row, column=3).number_format = '0.0%'
    tax_rate_row = row
    
    row += 1  # Row 35
    ws.cell(row=row, column=1, value='Tax on Original Gain (if paid)')
    ws.cell(row=row, column=3, value=f'=C{cap_gain_row}*C{tax_rate_row}')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    
    row += 2  # Row 37 - Traditional path header
    ws.cell(row=row, column=1, value='TRADITIONAL INVESTMENT PATH:')
    ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
    
    row += 1  # Row 38
    ws.cell(row=row, column=1, value='Amount Invested (after tax)')
    ws.cell(row=row, column=3, value=f'=C{cap_gain_row}-C35')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    trad_invest_row = row
    
    row += 1  # Row 39
    ws.cell(row=row, column=1, value='Return at Same Multiple')
    ws.cell(row=row, column=3, value=f'=C{trad_invest_row}*C29')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    
    row += 1  # Row 40
    ws.cell(row=row, column=1, value='Less: Capital Gains Tax (23.8%)')
    ws.cell(row=row, column=3, value=f'=-(C{row-1}-C{trad_invest_row})*0.238')
    ws.cell(row=row, column=3).number_format = '($#,##0)'
    
    row += 1  # Row 41
    ws.cell(row=row, column=1, value='Net After-Tax (Traditional)')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f'=C39+C40')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    trad_net_row = row
    
    row += 2  # Row 43 - OZ path header  
    ws.cell(row=row, column=1, value='OZ FUND PATH:')
    ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
    
    row += 1  # Row 44
    ws.cell(row=row, column=1, value='Amount Invested (full gain)')
    ws.cell(row=row, column=3, value=f'=C{cap_gain_row}')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    oz_invest_row = row
    
    row += 1  # Row 45
    ws.cell(row=row, column=1, value='Return at Same Multiple')
    ws.cell(row=row, column=3, value=f'=C{oz_invest_row}*C29')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    
    row += 1  # Row 46
    ws.cell(row=row, column=1, value='Less: Deferred Tax Due (2026)')
    ws.cell(row=row, column=3, value='=-C35')
    ws.cell(row=row, column=3).number_format = '($#,##0)'
    
    row += 1  # Row 47
    ws.cell(row=row, column=1, value='Tax on Appreciation (10yr hold)')
    ws.cell(row=row, column=3, value=0)  # Zero for 10-year hold
    ws.cell(row=row, column=3).number_format = '$#,##0'
    
    row += 1  # Row 48
    ws.cell(row=row, column=1, value='Net After-Tax (OZ Fund)')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value='=C45+C46+C47')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    oz_net_row = row
    
    row += 2  # Row 50
    ws.cell(row=row, column=1, value='OZ TAX SAVINGS')
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="006400")
    ws.cell(row=row, column=3, value=f'=C{oz_net_row}-C{trad_net_row}')
    ws.cell(row=row, column=3).number_format = '$#,##0'
    ws.cell(row=row, column=3).font = Font(bold=True, size=12, color="006400")
    format_total_row(ws, row, 4)
    
    # Column widths
    set_column_widths(ws, {'A': 32, 'B': 8, 'C': 18, 'D': 15})
    
    # Freeze panes
    ws.freeze_panes = 'A4'
    
    return ws

def main():
    """Build the complete OZ Fund Pro Forma workbook"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create all sheets
    sheets = [
        '1_Summary Dashboard',
        '2_Development Budget',
        '3_Unit Mix Revenue',
        '4_Operating Expenses',
        '5_Debt Schedule',
        '6_Cash Flow Analysis',
        '7_Returns Exit'
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(sheet_name)
    
    # Build each sheet
    print("Building Summary Dashboard...")
    build_summary_dashboard(wb)
    
    print("Building Development Budget...")
    build_development_budget(wb)
    
    print("Building Unit Mix & Revenue...")
    build_unit_mix_revenue(wb)
    
    print("Building Operating Expenses...")
    build_operating_expenses(wb)
    
    print("Building Debt Schedule...")
    build_debt_schedule(wb)
    
    print("Building Cash Flow Analysis...")
    build_cash_flow_analysis(wb)
    
    print("Building Returns & Exit Analysis...")
    build_returns_exit(wb)
    
    # Save workbook
    output_path = '/home/ubuntu/clawd/oz-fund-proforma-template.xlsx'
    wb.save(output_path)
    print(f"\n✅ Workbook saved to: {output_path}")
    
    # Verify
    from openpyxl import load_workbook
    wb_check = load_workbook(output_path)
    print(f"\nSheets created: {wb_check.sheetnames}")
    
    return output_path

if __name__ == '__main__':
    main()
